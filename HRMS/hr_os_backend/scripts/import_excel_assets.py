import os
import sys
import openpyxl
from uuid import uuid4
from datetime import datetime
from dotenv import load_dotenv

# Load env from backend
backend_dir = r"c:\Users\User\Downloads\HRMS (2)\HRMS\HRMS\hr_os_backend"
sys.path.insert(0, backend_dir)
os.chdir(backend_dir)
load_dotenv()

from app.core.mongodb import connect_mongo

def run():
    db = connect_mongo()
    print("Connected to database:", db.name)

    # 1. Clear existing IT assets
    print("Clearing existing IT assets...")
    db.it_assets.delete_many({})
    db["IT Assets"].delete_many({})

    # 2. Extract details
    entity_id = "Zipaworld"
    
    # Pre-fetch existing employees to build a mapping cache
    print("Building employee cache...")
    employees = list(db.employees.find({}))
    users = list(db.users.find({}))
    
    # Map name to employee
    emp_by_name = {}
    for emp in employees:
        emp_by_name[emp["full_name"].strip().lower()] = emp
        
    emp_by_email = {}
    for emp in employees:
        if emp.get("email"):
            emp_by_email[emp["email"].strip().lower()] = emp

    # Default password hash for "password123"
    # We will reuse this hash for all newly created users
    default_password_hash = "$2b$12$f3Dyvsc/FpWqHF4f.MtTgOx0b3eDFfG105Oz9QTdBbCIuMmGO.dW6"

    # Helper function to find or create employee & user
    def get_or_create_employee(name_str, email_str=None):
        if not name_str:
            return None, ""
            
        name_clean = name_str.strip()
        name_lower = name_clean.lower()
        
        # 1. Try to match by email if provided
        if email_str:
            email_clean = email_str.strip().lower()
            if email_clean in emp_by_email:
                emp = emp_by_email[email_clean]
                return emp["_id"], emp["full_name"]
                
        # 2. Try to match by name
        if name_lower in emp_by_name:
            emp = emp_by_name[name_lower]
            return emp["_id"], emp["full_name"]
            
        # Try substring match
        for key, emp in emp_by_name.items():
            if name_lower in key or key in name_lower:
                return emp["_id"], emp["full_name"]
                
        # 3. Not found, create new employee & user
        print(f"Creating new user/employee for: {name_clean}")
        uid = str(uuid4())
        eid = str(uuid4())
        
        # Resolve email
        if email_str:
            email = email_str.strip().lower()
        else:
            email_part = name_lower.replace(" ", ".")
            email = f"{email_part}@zipaworld.com"
            
        # Check if username or email already exists in users collection to prevent duplicate keys
        if db.users.find_one({"email": email}):
            # Append a random prefix to prevent collision
            email = f"{email.split('@')[0]}_{str(uuid4())[:4]}@zipaworld.com"

        # Create user
        user_doc = {
            "_id": uid,
            "id": uid,
            "email": email,
            "hashed_password": default_password_hash,
            "password_hash": default_password_hash,
            "role": "EMPLOYEE",
            "entity_id": entity_id,
            "employee_id": eid,
        }
        db.users.insert_one(user_doc)
        
        # Create employee
        emp_doc = {
            "_id": eid,
            "id": eid,
            "user_id": uid,
            "entity_id": entity_id,
            "employee_code": f"EXP-USR-{str(uuid4())[:6].upper()}",
            "full_name": name_clean,
            "email": email,
            "designation": "Staff",
            "status": "ACTIVE",
            "joining_date": datetime.utcnow(),
            "personal_info": {
                "phone": "+919999999999",
                "gender": "Male",
                "marital_status": "Single"
            },
            "employment_details": {
                "attendance_mode": "BIOMETRIC",
                "work_location": "Office"
            },
            "salary_structure": {
                "basic": 30000,
                "hra": 12000,
                "allowances": 8000
            }
        }
        db.employees.insert_one(emp_doc)
        
        # Cache it
        emp_by_name[name_lower] = emp_doc
        emp_by_email[email] = emp_doc
        
        return eid, name_clean

    def safe_get(row, idx, default=None):
        if row is None or idx >= len(row):
            return default
        return row[idx]

    # 3. Read Laptop Register
    laptop_file = r"C:\Users\User\Downloads\Asset Laptop (1).xlsx"
    print(f"Reading laptop file: {laptop_file}")
    
    wb_laptop = openpyxl.load_workbook(laptop_file, read_only=True, data_only=True)
    ws_laptop = wb_laptop["Asset Register"]
    
    current_location = "Office"
    current_asset_type = "laptop"
    laptops_count = 0

    # Start scanning from Row 6
    rows = list(ws_laptop.iter_rows(values_only=True))
    for i, r in enumerate(rows[5:], start=6):
        # Check if it's a separator or category row
        row_str = " ".join([str(val) for val in r if val is not None]).strip()
        if not row_str:
            continue
            
        # Check for section boundaries
        non_empty_cells = [val for val in r if val is not None]
        if len(non_empty_cells) == 1:
            val = non_empty_cells[0]
            val_clean = str(val).strip()
            if "Warehouse" in val_clean:
                current_location = "Warehouse"
                current_asset_type = "laptop"
                print(f"Row {i}: Section switched to Warehouse")
                continue
            elif "Office" in val_clean:
                current_location = "Office"
                current_asset_type = "laptop"
                print(f"Row {i}: Section switched to Office")
                continue
            elif "Mobile Phones" in val_clean:
                current_location = "Office"
                current_asset_type = "mobile_phone"
                print(f"Row {i}: Section switched to Mobile Phones")
                continue
                
        # Skip header duplicates or empty entries
        brand = safe_get(r, 2)
        serial = safe_get(r, 5)
        model = safe_get(r, 4)
        if not brand and not serial and not model:
            continue
            
        assigned_to_name = safe_get(r, 0)
        email_office = safe_get(r, 1)
        
        assigned_to_id = None
        assigned_to_name_resolved = ""
        
        if assigned_to_name and str(assigned_to_name).strip().lower() not in ["warehouse", "office", "available", "none", "n/a"]:
            assigned_to_id, assigned_to_name_resolved = get_or_create_employee(assigned_to_name, email_office)
            status = "ASSIGNED"
        else:
            status = "AVAILABLE"

        # Build asset doc
        asset_id = str(uuid4())
        raw_asset_type = safe_get(r, 10)
        raw_location = safe_get(r, 11)
        
        asset_doc = {
            "_id": asset_id,
            "id": asset_id,
            "entity_id": entity_id,
            "asset_type": current_asset_type if raw_asset_type is None else str(raw_asset_type).strip().lower(),
            "asset_tag": str(serial).strip() if serial else f"SN-{str(uuid4())[:8].upper()}",
            "brand": str(brand).strip() if brand else "",
            "desktop_name": str(safe_get(r, 3)).strip() if safe_get(r, 3) else "",
            "model": str(model).strip() if model else "",
            "processor": str(safe_get(r, 6)).strip() if safe_get(r, 6) else "",
            "ram": str(safe_get(r, 7)).strip() if safe_get(r, 7) else "",
            "storage": str(safe_get(r, 8)).strip() if safe_get(r, 8) else "",
            "operating_system": str(safe_get(r, 9)).strip() if safe_get(r, 9) else "",
            "location": current_location if raw_location is None else str(raw_location).strip(),
            "issue": str(safe_get(r, 13)).strip() if safe_get(r, 13) else "",
            "gpu": str(safe_get(r, 14)).strip() if safe_get(r, 14) else "",
            "assigned_to": assigned_to_id,
            "assigned_to_name": assigned_to_name_resolved,
            "status": status,
            "accessories": str(safe_get(r, 16)).strip() if safe_get(r, 16) else "",
            "created_at": datetime.utcnow()
        }
        db.it_assets.insert_one(asset_doc)
        laptops_count += 1

    print(f"Imported {laptops_count} laptops/mobile phones.")

    # 4. Read Stock Management
    stock_file = r"C:\Users\User\Downloads\Asset_Stock_Management (1).xlsx"
    print(f"Reading stock file: {stock_file}")
    
    wb_stock = openpyxl.load_workbook(stock_file, read_only=True, data_only=True)
    ws_stock = wb_stock["Current Stock"]
    
    stock_issued_count = 0
    stock_available_count = 0

    rows_stock = list(ws_stock.iter_rows(values_only=True))
    for i, r in enumerate(rows_stock[1:], start=2):
        category = safe_get(r, 0)
        item_name = safe_get(r, 1)
        brand_model = safe_get(r, 2)
        
        if not category and not item_name:
            continue
            
        issued_to_str = safe_get(r, 6)
        remaining_stock = safe_get(r, 5)
        
        # Parse issued users
        issued_names = []
        if issued_to_str:
            # Splitting by comma
            issued_names = [name.strip() for name in str(issued_to_str).split(",") if name.strip()]
            
        # Create assigned assets for each issued person
        for name in issued_names:
            assigned_to_id, assigned_to_name_resolved = get_or_create_employee(name)
            
            asset_id = str(uuid4())
            asset_doc = {
                "_id": asset_id,
                "id": asset_id,
                "entity_id": entity_id,
                "asset_type": str(category).strip().lower() if category else "peripheral",
                "asset_tag": f"STK-ISS-{str(item_name).replace(' ', '').upper()}-{name.replace(' ', '').upper()[:5]}-{str(uuid4())[:4].upper()}",
                "brand": str(brand_model).strip() if brand_model else "",
                "desktop_name": f"{item_name} ({name})",
                "model": str(item_name).strip() if item_name else "",
                "processor": "",
                "ram": "",
                "storage": "",
                "operating_system": "",
                "location": "Office",
                "issue": "",
                "gpu": "",
                "assigned_to": assigned_to_id,
                "assigned_to_name": assigned_to_name_resolved,
                "status": "ASSIGNED",
                "accessories": "",
                "created_at": datetime.utcnow()
            }
            db.it_assets.insert_one(asset_doc)
            stock_issued_count += 1
            
        # Create AVAILABLE assets for remaining quantity
        remaining_count = 0
        if remaining_stock:
            try:
                # Handle standard float or int conversion
                remaining_count = int(float(remaining_stock))
            except ValueError:
                # If remaining stock has text like "9 unused", parse the integer
                try:
                    num_parts = [int(s) for s in str(remaining_stock).split() if s.isdigit()]
                    if num_parts:
                        remaining_count = num_parts[0]
                except Exception:
                    remaining_count = 0
                    
        for idx in range(1, remaining_count + 1):
            asset_id = str(uuid4())
            asset_doc = {
                "_id": asset_id,
                "id": asset_id,
                "entity_id": entity_id,
                "asset_type": str(category).strip().lower() if category else "peripheral",
                "asset_tag": f"STK-AVL-{str(item_name).replace(' ', '').upper()}-{idx:02d}-{str(uuid4())[:4].upper()}",
                "brand": str(brand_model).strip() if brand_model else "",
                "desktop_name": str(item_name).strip(),
                "model": str(item_name).strip() if item_name else "",
                "processor": "",
                "ram": "",
                "storage": "",
                "operating_system": "",
                "location": "Warehouse",
                "issue": "",
                "gpu": "",
                "assigned_to": None,
                "assigned_to_name": "",
                "status": "AVAILABLE",
                "accessories": "",
                "created_at": datetime.utcnow()
            }
            db.it_assets.insert_one(asset_doc)
            stock_available_count += 1

    print(f"Imported {stock_issued_count} issued stock items and {stock_available_count} available stock items.")

    # 5. Read Stock Management rows directly into it_stock collection for aggregated view
    print("Seeding aggregated it_stock collection...")
    db.it_stock.delete_many({})
    stock_count = 0
    for i, r in enumerate(rows_stock[1:], start=2):
        category = safe_get(r, 0)
        item_name = safe_get(r, 1)
        brand_model = safe_get(r, 2)
        
        if not category and not item_name:
            continue
            
        total_stock = safe_get(r, 3)
        issued_qty = safe_get(r, 4)
        remaining_stock = safe_get(r, 5)
        issued_to = safe_get(r, 6)
        department = safe_get(r, 7)
        issue_date = safe_get(r, 8)
        unit = safe_get(r, 9)
        reorder = safe_get(r, 10)
        
        date_str = ""
        if issue_date:
            if hasattr(issue_date, "strftime"):
                date_str = issue_date.strftime("%Y-%m-%d")
            else:
                date_str = str(issue_date)
                
        stock_id = str(uuid4())
        stock_doc = {
            "_id": stock_id,
            "id": stock_id,
            "entity_id": entity_id,
            "category": str(category).strip() if category else "",
            "item_name": str(item_name).strip() if item_name else "",
            "brand_model": str(brand_model).strip() if brand_model else "",
            "total_stock": str(total_stock).strip() if total_stock is not None else "",
            "issued_qty": str(issued_qty).strip() if issued_qty is not None else "",
            "remaining_qty": str(remaining_stock).strip() if remaining_stock is not None else "",
            "issued_to": str(issued_to).strip() if issued_to else "",
            "department": str(department).strip() if department else "",
            "issue_date": date_str,
            "unit": str(unit).strip() if unit else "",
            "reorder_required": str(reorder).strip() if reorder else "No",
            "created_at": datetime.utcnow()
        }
        db.it_stock.insert_one(stock_doc)
        stock_count += 1
        
    print(f"Imported {stock_count} original stock items into it_stock collection.")
    print("DB seeding of Excel assets complete!")

if __name__ == "__main__":
    run()
